import openpyxl


class ReadAcess:
    def __init__(self, caminho_arq: str = ''):
        self.caminho_arq = caminho_arq
        self.dados = None

    def process_arq(self, action):
        with open(f'{self.caminho_arq}{action}.txt', 'r') as arq:
            self.dados = arq.readlines()
            self.dados = [linha.replace('\n', '').split(';')
                          for linha in self.dados]

    def get_dates(self):
        return self.dados


class manager:
    def __init__(self):
        self.work = openpyxl.Workbook()
        self.planilha_ativa = None

    def active(self):
        self.planilha_ativa = self.work.active

    def add(self, title: str = ''):

        nova_planilha = self.work.create_sheet(title)
        self.work.active = nova_planilha
        self.planilha_ativa = nova_planilha

        return nova_planilha

    def add_header(self, dados: list):
        self.planilha_ativa.append(dados)

    def update_date(self, celula: str, dado):
        self.planilha_ativa[celula] = dado

    def cells_size(self, celula: str, size: int):
        self.planilha_ativa.column_dimensions[celula].width = size

    def column_alignment(self):
        for ws_name in self.planilha_ativa.iter_rows():
            for cell in ws_name:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal='center', vertical='center')

    def merge_date(self, cel_start: str, cel_end: str):
        self.planilha_ativa.merge_cells(f'{cel_start}:{cel_end}')

    def apply_styles(self, cel: str, styles: list):
        for style in styles:
            setattr(self.planilha_ativa[cel], style[0], style[1])

    def add_graph(self, celula: str, comprimento: float, height: float,
                  title: str, title_x: str, title_y: str,
                  ref_x: openpyxl.chart.Reference, ref_y:
                  openpyxl.chart.Reference, GraphPriority: list):

        graph = openpyxl.chart.LineChart()
        graph.width = comprimento
        graph.height = height
        graph.title = title
        graph.x_axis.title = title_x
        graph.y_axis.title = title_y

        graph.add_data(ref_x)
        graph.set_categories(ref_y)

        for serie, priority in zip(graph.series, GraphPriority):
            serie.graphicalProperties.line.width = priority.thick
            serie.graphicalProperties.line.solidFill = priority.color

        self.planilha_ativa.add_chart(graph, celula)

    def add_img(self, celula: str, path: str):
        img = openpyxl.drawing.image.Image(path)
        self.planilha_ativa.add_image(img, celula)

    def save(self, path: str):
        self.work.save(path)


class SeriesGraphPriority:
    def __init__(self, thick: int, color: str):
        self.thick = thick
        self.color = color
